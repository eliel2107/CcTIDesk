import os, uuid, csv, json
from io import BytesIO
from datetime import datetime
from .notify import notify_ticket_created, notify_ticket_assigned, notify_status_changed
from .models import list_tickets_paginated
from app.extensions import limiter
from io import StringIO
from flask import Blueprint, render_template, request, redirect, url_for, flash, Response, current_app, send_from_directory, abort, g, jsonify, send_file
from werkzeug.utils import secure_filename

from .db import init_db, get_db
from . import ai_service
from .services.ai.gemini_client import GeminiClientError
from .auth import login_required, role_required
from .address_book import DELIVERY_PRESETS, SENDER_PRESETS
from .services.asset_service import list_assets_for_select, get_asset
from .notifications import (
    on_ticket_criado, on_ticket_assumido, on_status_atualizado,
    on_aguardando_confirmacao, on_conclusao_rejeitada, on_chamado_concluido,
    on_aprovacao_necessaria, on_chamado_aprovado, on_chamado_reprovado, on_transferencia,
    get_notificacoes, contar_nao_lidas, marcar_lida, marcar_todas_lidas,
)

from .models import (
    STATUSES, TYPES, PRIORITIES, CLASSIFICATIONS,
    _today_ymd,
    create_ticket, list_tickets, get_ticket, update_status, update_fields, get_logs,
    finalizar_ticket, confirmar_conclusao, rejeitar_conclusao,
    dashboard_stats, dashboard_stats_advanced, list_attachments, add_attachment, count_attachments,
    list_steps, toggle_step, add_step, delete_step, move_step,
    delete_attachment,
    list_tickets_by_requester, list_queue_tickets, assign_ticket,
    list_categories, get_user_categories,
    # Novas melhorias
    add_comment, get_comments, delete_comment,
    transfer_ticket, get_transfers,
    reabrir_ticket,
    devolver_ao_solicitante, reenviar_pelo_solicitante,
    aprovar_ticket, reprovar_ticket,
    get_sla_status,
    tma_stats,
    list_groups, get_group, create_group, update_group,
    set_group_members, set_group_categories, get_group_members, get_group_categories,
    assign_ticket_to_group,
    list_webhooks, create_webhook, update_webhook, delete_webhook,
    search_tickets_advanced,
    create_category_full, update_category_full,
)

bp = Blueprint("routes", __name__)


def _build_logs_filters_and_query(args):
    filtros = {
        "q": (args.get("q") or "").strip(),
        "evento": (args.get("evento") or "").strip(),
        "ticket_id": (args.get("ticket_id") or "").strip(),
        "data_inicio": (args.get("data_inicio") or "").strip(),
        "data_fim": (args.get("data_fim") or "").strip(),
    }

    where = []
    params = []

    if filtros["ticket_id"].isdigit():
        where.append("tl.ticket_id = ?")
        params.append(int(filtros["ticket_id"]))

    if filtros["evento"]:
        where.append("tl.evento = ?")
        params.append(filtros["evento"].upper())

    if filtros["q"]:
        like = f"%{filtros['q']}%"
        where.append("(tl.detalhe LIKE ? OR tl.evento LIKE ? OR t.titulo LIKE ? OR t.numero_chamado LIKE ?)")
        params.extend([like, like, like, like])

    if filtros["data_inicio"]:
        where.append("date(tl.criado_em) >= date(?)")
        params.append(filtros["data_inicio"])

    if filtros["data_fim"]:
        where.append("date(tl.criado_em) <= date(?)")
        params.append(filtros["data_fim"])

    where_sql = f"WHERE {' AND '.join(where)}" if where else ""
    return filtros, where_sql, params


def _fetch_logs_for_audit(db, where_sql, params, limit=500):
    sql = f"""
        SELECT tl.id, tl.ticket_id, tl.evento, tl.detalhe, tl.criado_em,
               t.numero_chamado, t.titulo, t.status
          FROM ticket_log tl
          LEFT JOIN tickets t ON t.id = tl.ticket_id
          {where_sql}
         ORDER BY tl.criado_em DESC, tl.id DESC
    """
    if limit:
        sql += f" LIMIT {int(limit)}"
    return db.execute(sql, params).fetchall()


def _logs_export_filename(prefix, ext):
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{stamp}.{ext}"


@bp.before_app_request
def ensure_db():
    try:
        init_db()
    except Exception:
        pass

def _allowed_file(filename: str) -> bool:
    if "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in current_app.config["ALLOWED_EXTENSIONS"]

def _solicitante_can_edit(ticket) -> bool:
    """Retorna True se o usuário atual pode editar o chamado.
    Solicitantes ficam bloqueados enquanto houver um responsável atribuído.
    Quando assigned_user_id for removido (None), o acesso é restaurado.
    """
    if g.user and g.user["role"] == "solicitante":
        if ticket and ticket["assigned_user_id"]:
            return False
    return True

@bp.get("/search")
@login_required
def global_search():
    q = (request.args.get("q") or "").strip()
    db = get_db()
    tickets = []
    assets = []
    if q:
        like = f"%{q}%"
        tickets = db.execute("""SELECT id, tipo, titulo, responsavel, prioridade, status, atualizado_em
                                FROM tickets
                                WHERE titulo LIKE ? OR solicitante LIKE ? OR responsavel LIKE ? OR codigo_rastreio LIKE ? OR descricao LIKE ?
                                ORDER BY atualizado_em DESC LIMIT 20""", (like, like, like, like, like)).fetchall()
        assets = db.execute("""SELECT id, tag, tipo, modelo, serial_number, local_base, responsavel, status, atualizado_em
                               FROM assets
                               WHERE tag LIKE ? OR modelo LIKE ? OR serial_number LIKE ? OR local_base LIKE ? OR responsavel LIKE ?
                               ORDER BY atualizado_em DESC LIMIT 20""", (like, like, like, like, like)).fetchall()
    return render_template("search.html", q=q, tickets=tickets, assets=assets)


@bp.get("/novo-chamado")
@login_required
@role_required('solicitante')
def requester_new_ticket():
    import json as _json
    categories = list_categories(only_active=True)
    cat_id = request.args.get("cat", "")
    selected_cat = None
    if cat_id:
        try:
            selected_cat = next((c for c in categories if str(c["id"]) == cat_id), None)
        except Exception:
            pass
    classif = request.args.get("classif", "REQUISICAO")
    if not cat_id or not selected_cat:
        return render_template("ticket_category_picker.html", categories=categories, next_url=url_for("routes.requester_new_ticket"))
    campos = []
    try:
        campos = _json.loads(selected_cat["campos_visiveis"] or "[]")
    except Exception:
        pass
    return render_template("requester_new_ticket.html", TYPES=TYPES, PRIORITIES=PRIORITIES, CLASSIFICATIONS=CLASSIFICATIONS,
                           categories=categories, selected_cat=selected_cat, campos_visiveis=campos, classif=classif,
                           ai_enabled=current_app.config.get("AI_ASSIST_ENABLED", False))

@bp.post("/novo-chamado")
@login_required
@role_required('solicitante')
@limiter.limit("20 per hour", error_message="Limite de chamados atingido. Aguarde antes de abrir mais chamados.")
def requester_create_ticket():
    try:
        data = dict(request.form)
        data["requester_user_id"] = g.user["id"]
        data["solicitante"] = g.user["nome"]
        tid = create_ticket(data)
        _t = get_ticket(tid)
        notify_ticket_created(
            current_app.config,
            ticket_id=tid,
            titulo=data.get("titulo",""),
            solicitante=g.user["nome"],
            tipo=data.get("tipo",""),
        )
        on_ticket_criado(tid, data.get("titulo",""), _t["categoria_id"] if _t else None, g.user["id"])
        if _t and _t["status"] == "AGUARDANDO_APROVACAO":
            on_aprovacao_necessaria(tid, data.get("titulo",""))
            flash(f"Chamado criado e aguardando aprovação de um admin.", "info")
        else:
            flash(f"Chamado criado com ID {tid}.", "success")
        return redirect(url_for("routes.my_tickets"))
    except Exception as e:
        flash(str(e), "error")
        return redirect(url_for("routes.requester_new_ticket"))

@bp.get("/meus-chamados")
@login_required
@role_required('solicitante')
def my_tickets():
    tickets = list_tickets_by_requester(g.user["id"])
    return render_template("my_tickets.html", tickets=tickets, stats_today=_today_ymd())

@bp.get("/fila")
@login_required
@role_required('admin', 'operador')
def queue():
    filters = {k: request.args.get(k, "") for k in ["status","tipo","q","only_unassigned","categoria_id"]}
    tickets = list_queue_tickets(
        filters,
        user_id=g.user["id"],
        user_role=g.user["role"],
    )
    categories = list_categories(only_active=True)
    user_cat_ids = get_user_categories(g.user["id"]) if g.user["role"] == "operador" else []
    return render_template("queue.html", tickets=tickets, STATUSES=STATUSES, TYPES=TYPES,
                           filters=filters, categories=categories, user_cat_ids=user_cat_ids,
                           assignment_timeout_minutes=current_app.config.get("ASSIGNMENT_TIMEOUT_MINUTES", 15))

@bp.post("/fila/<int:ticket_id>/assumir")
@login_required
@role_required('admin', 'operador')
def take_ticket(ticket_id: int):
    try:
        assign_ticket(ticket_id, g.user["id"], g.user["nome"])
        _t = get_ticket(ticket_id)
        if _t:
            from .services.auth_service import get_user
            _req = get_user(_t["requester_user_id"]) if _t["requester_user_id"] else None
            notify_ticket_assigned(
                current_app.config,
                ticket_id=ticket_id,
                titulo=_t["titulo"],
                responsavel=g.user["nome"],
                requester_email=_req["email"] if _req else None,
            )
            if _t["requester_user_id"]:
                on_ticket_assumido(ticket_id, _t["titulo"], _t["requester_user_id"], g.user["nome"])
        flash("Chamado assumido. Iniciando atendimento.", "success")
    except Exception as e:
        flash(str(e), "error")
        return redirect(url_for("routes.queue"))
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

@bp.get("/api/address-book")

@login_required
def api_address_book():
    return {"delivery": DELIVERY_PRESETS, "sender": SENDER_PRESETS}

@bp.get("/api/notificacoes")
@login_required
def api_notificacoes():
    from flask import jsonify
    notifs = get_notificacoes(g.user["id"], limite=20)
    nao_lidas = contar_nao_lidas(g.user["id"])
    return jsonify({
        "nao_lidas": nao_lidas,
        "notificacoes": [
            {
                "id": n["id"],
                "tipo": n["tipo"],
                "titulo": n["titulo"],
                "mensagem": n["mensagem"],
                "ticket_id": n["ticket_id"],
                "lida": bool(n["lida"]),
                "criado_em": n["criado_em"],
            } for n in notifs
        ]
    })

@bp.post("/api/notificacoes/<int:notif_id>/ler")
@login_required
def api_marcar_lida(notif_id: int):
    from flask import jsonify
    marcar_lida(notif_id, g.user["id"])
    return jsonify({"ok": True})

@bp.post("/api/notificacoes/ler-todas")
@login_required
def api_marcar_todas_lidas():
    from flask import jsonify
    marcar_todas_lidas(g.user["id"])
    return jsonify({"ok": True})



@bp.post("/api/ai/opening-assistant")
@login_required
def api_ai_opening_assistant():
    data = request.get_json(silent=True) or {}
    try:
        result = ai_service.opening_assistant(
            description=(data.get('descricao') or '').strip(),
            title=(data.get('titulo') or '').strip(),
            category=(data.get('categoria') or '').strip(),
        )
        return jsonify(result)
    except GeminiClientError as e:
        return jsonify({'enabled': True, 'error': str(e)}), 502
    except Exception as e:
        current_app.logger.exception('Falha no assistente de abertura')
        return jsonify({'enabled': True, 'error': f'Falha ao consultar IA: {e}'}), 500


@bp.post("/api/ai/tickets/<int:ticket_id>/resolution-draft")
@login_required
@role_required('admin', 'operador')
def api_ai_resolution_draft(ticket_id: int):
    t = get_ticket(ticket_id)
    if not t:
        return jsonify({'error': 'Chamado não encontrado.'}), 404
    data = request.get_json(silent=True) or {}
    resolution = (data.get('resolution') or '').strip()
    if not resolution:
        return jsonify({'error': 'Informe a resolução aplicada.'}), 400
    try:
        result = ai_service.resolution_assistant(
            ticket_id=ticket_id,
            resolution_text=resolution,
            title=t['titulo'] or '',
            description=t['descricao'] or '',
        )
        return jsonify(result)
    except GeminiClientError as e:
        return jsonify({'enabled': True, 'error': str(e)}), 502
    except Exception as e:
        current_app.logger.exception('Falha ao gerar rascunho de resolução')
        return jsonify({'enabled': True, 'error': f'Falha ao consultar IA: {e}'}), 500

@bp.get("/dashboard")
@login_required
def dashboard():
    if g.user and g.user["role"] == "solicitante":
        return redirect(url_for("routes.my_tickets"))
    stats = dashboard_stats_advanced(
        user_id=g.user["id"] if g.user else None,
        user_role=g.user["role"] if g.user else None,
    )
    return render_template("dashboard.html", stats=stats, STATUSES=STATUSES,
                           user_role=g.user["role"] if g.user else None,
                           user_nome=g.user["nome"] if g.user else None)

@bp.get("/home")
@login_required
def executive_home():
    db = get_db()
    open_count = db.execute("SELECT COUNT(*) as c FROM tickets WHERE status NOT IN ('CONCLUIDO','CANCELADO')").fetchone()["c"]
    overdue_count = db.execute("SELECT COUNT(*) as c FROM tickets WHERE data_limite IS NOT NULL AND data_limite < date('now') AND status NOT IN ('CONCLUIDO','CANCELADO')").fetchone()["c"]
    asset_total = db.execute("SELECT COUNT(*) as c FROM assets").fetchone()["c"] if db.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='assets'").fetchone() else 0
    assets_in_transit = db.execute("SELECT COUNT(*) as c FROM assets WHERE status='EM_TRANSITO'").fetchone()["c"] if asset_total >= 0 else 0
    assets_without_owner = db.execute("SELECT COUNT(*) as c FROM assets WHERE COALESCE(TRIM(responsavel),'')=''").fetchone()["c"] if asset_total >= 0 else 0
    return render_template("home_exec.html", open_count=open_count, overdue_count=overdue_count, asset_total=asset_total, assets_in_transit=assets_in_transit, assets_without_owner=assets_without_owner)

@bp.get("/")
@login_required
def index():

    filters = {k: request.args.get(k, "") for k in ["status","tipo","prioridade","responsavel","q","sort_by","asset_id","show_archived"]}
    try:
        page = max(1, int(request.args.get("page", 1)))
    except (ValueError, TypeError):
        page = 1
    per_page = 25
    tickets, total, total_pages = list_tickets_paginated(
        filters, page=page, per_page=per_page,
        user_id=g.user["id"] if g.user else None,
        user_role=g.user["role"] if g.user else None,
    )
    grouped_tickets = []
    if g.user and g.user["role"] == "admin":
        grouped = {}
        for t in tickets:
            grouped.setdefault(t["categoria_nome"] or "Sem categoria", []).append(t)
        grouped_tickets = [{"categoria": k, "tickets": v} for k, v in grouped.items()]
    stats_today = _today_ymd()
    return render_template(
        "index.html",
        tickets=tickets, grouped_tickets=grouped_tickets,
        STATUSES=STATUSES, TYPES=TYPES, PRIORITIES=PRIORITIES,
        filters=filters, stats_today=stats_today,
        page=page, total=total, total_pages=total_pages, per_page=per_page,
        user_role=g.user["role"] if g.user else None,
    )

@bp.get("/tickets/new")
@login_required
@role_required('admin', 'operador')
def new_ticket():
    import json as _json
    categories = list_categories(only_active=True)
    cat_id = request.args.get("cat", "")
    selected_cat = None
    if cat_id:
        try:
            selected_cat = next((c for c in categories if str(c["id"]) == cat_id), None)
        except Exception:
            pass
    classif = request.args.get("classif", "REQUISICAO")
    if not cat_id or not selected_cat:
        # Step 1: category picker
        return render_template("ticket_category_picker.html", categories=categories, next_url=url_for("routes.new_ticket"))
    # Step 2: full form with pre-selected category
    campos = []
    try:
        campos = _json.loads(selected_cat["campos_visiveis"] or "[]")
    except Exception:
        pass
    return render_template("new_ticket.html", TYPES=TYPES, PRIORITIES=PRIORITIES, CLASSIFICATIONS=CLASSIFICATIONS,
                           assets=list_assets_for_select(), categories=categories,
                           selected_cat=selected_cat, campos_visiveis=campos, classif=classif)

@bp.post("/tickets")
@login_required
@role_required('admin', 'operador')
@limiter.limit("60 per hour", error_message="Limite de criação de chamados atingido.")
def create_ticket_route():
    try:
        data2 = dict(request.form)
        # Garante que quem abriu o chamado fica registrado como solicitante
        if not data2.get("requester_user_id"):
            data2["requester_user_id"] = str(g.user["id"])
        tid = create_ticket(data2)
        _t = get_ticket(tid)
        notify_ticket_created(
            current_app.config,
            ticket_id=tid,
            titulo=data2.get("titulo",""),
            solicitante=data2.get("solicitante", g.user["nome"]),
            tipo=data2.get("tipo",""),
        )
        on_ticket_criado(tid, data2.get("titulo",""), _t["categoria_id"] if _t else None, g.user["id"])
        if _t and _t["status"] == "AGUARDANDO_APROVACAO":
            on_aprovacao_necessaria(tid, data2.get("titulo",""))
            flash(f"Chamado criado e aguardando aprovação.", "info")
        else:
            flash(f"Chamado criado com ID {tid}.", "success")
        return redirect(url_for("routes.ticket_detail", ticket_id=tid))
    except Exception as e:
        flash(str(e), "error")
        return redirect(url_for("routes.new_ticket"))

@bp.get("/tickets/<int:ticket_id>")
@login_required
def ticket_detail(ticket_id: int):
    t = get_ticket(ticket_id)
    if g.user and g.user["role"] == "solicitante" and t and t["requester_user_id"] != g.user["id"]:
        flash("Você não tem permissão para acessar este chamado.", "error")
        return redirect(url_for("routes.my_tickets"))
    if not t:
        flash("Chamado não encontrado.", "error")
        return redirect(url_for("routes.index"))
    logs = get_logs(ticket_id)
    attachments = list_attachments(ticket_id)
    steps = list_steps(ticket_id)
    comments = get_comments(ticket_id, include_internal=(g.user["role"] in ("admin","operador")))
    transfers = get_transfers(ticket_id)
    max_files = current_app.config["MAX_ATTACHMENTS_PER_TYPE"].get(t["tipo"], 20)
    allowed = ", ".join(sorted(current_app.config["ALLOWED_EXTENSIONS"]))
    asset = get_asset(t["asset_id"]) if "asset_id" in t.keys() and t["asset_id"] else None
    from .models import get_category
    categoria = get_category(t["categoria_id"]) if t["categoria_id"] else None
    sla_status = get_sla_status(t)
    # Lista de operadores para transferência
    operators = []
    if g.user["role"] in ("admin", "operador"):
        db = get_db()
        operators = db.execute(
            "SELECT id, nome FROM users WHERE active=1 AND role IN ('admin','operador') AND id!=? ORDER BY nome",
            (g.user["id"],)
        ).fetchall()
    groups = list_groups() if g.user["role"] in ("admin","operador") else []
    # Artigos KB relacionados
    from .kb import suggest_articles
    from .services.stock_service import consumos_por_ticket, produtos_para_select
    kb_articles = suggest_articles(t["titulo"], t["categoria_id"]) if t["titulo"] else []
    consumos_ticket = consumos_por_ticket(ticket_id)
    produtos_estoque = produtos_para_select() if g.user["role"] in ("admin","operador") else []
    return render_template(
        "ticket_detail.html",
        ticket=t,
        asset=asset,
        categoria=categoria,
        logs=logs,
        attachments=attachments,
        steps=steps,
        comments=comments,
        transfers=transfers,
        sla_status=sla_status,
        operators=operators,
        groups=groups,
        kb_articles=kb_articles,
        consumos_ticket=consumos_ticket,
        produtos_estoque=produtos_estoque,
        STATUSES=STATUSES,
        max_files=max_files,
        allowed_ext=allowed,
        assets=list_assets_for_select(),
        latest_ai_insight=ai_service.last_ticket_insight(ticket_id),
        ai_enabled=current_app.config.get("AI_ASSIST_ENABLED", False)
    )

@bp.post("/tickets/<int:ticket_id>/status")
@login_required
@role_required('admin', 'operador')
def update_status_route(ticket_id: int):
    try:
        new_st = request.form.get("status","")
        update_status(ticket_id, new_st, request.form.get("detalhe",""))
        _t = get_ticket(ticket_id)
        if _t:
            from .services.auth_service import get_user
            _req = get_user(_t["requester_user_id"]) if _t["requester_user_id"] else None
            notify_status_changed(
                current_app.config,
                ticket_id=ticket_id,
                titulo=_t["titulo"],
                new_status=new_st,
                requester_email=_req["email"] if _req else None,
            )
            if _t["requester_user_id"]:
                on_status_atualizado(ticket_id, _t["titulo"], _t["requester_user_id"], new_st)
        flash("Status atualizado.", "success")
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

@bp.post("/tickets/<int:ticket_id>/edit")
@login_required
@role_required('admin', 'operador')
def edit_ticket_route(ticket_id: int):
    try:
        update_fields(ticket_id, dict(request.form))
        flash("Chamado atualizado.", "success")
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

@bp.post("/tickets/<int:ticket_id>/finalizar")
@login_required
@role_required('admin', 'operador')
def finalizar_ticket_route(ticket_id: int):
    try:
        _t = get_ticket(ticket_id)
        finalizar_ticket(ticket_id, g.user["nome"])
        if _t and _t["requester_user_id"]:
            on_aguardando_confirmacao(ticket_id, _t["titulo"], _t["requester_user_id"], g.user["nome"])
        flash("Chamado enviado para confirmação de quem abriu.", "success")
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

@bp.post("/tickets/<int:ticket_id>/confirmar")
@login_required
def confirmar_ticket_route(ticket_id: int):
    t = get_ticket(ticket_id)
    if not t:
        flash("Chamado não encontrado.", "error")
        return redirect(url_for("routes.index"))
    if g.user["role"] not in ("admin",) and t["requester_user_id"] != g.user["id"]:
        flash("Apenas quem abriu o chamado pode confirmar a conclusão.", "error")
        return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))
    try:
        confirmar_conclusao(ticket_id, g.user["nome"])
        on_chamado_concluido(ticket_id, t["titulo"], t["requester_user_id"], t["assigned_user_id"], g.user["nome"])
        flash("Chamado concluído com sucesso!", "success")
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

@bp.post("/tickets/<int:ticket_id>/rejeitar")
@login_required
def rejeitar_ticket_route(ticket_id: int):
    t = get_ticket(ticket_id)
    if not t:
        flash("Chamado não encontrado.", "error")
        return redirect(url_for("routes.index"))
    if g.user["role"] not in ("admin",) and t["requester_user_id"] != g.user["id"]:
        flash("Apenas quem abriu o chamado pode rejeitar a conclusão.", "error")
        return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))
    try:
        motivo = request.form.get("motivo", "")
        rejeitar_conclusao(ticket_id, motivo, g.user["nome"])
        if t["assigned_user_id"]:
            on_conclusao_rejeitada(ticket_id, t["titulo"], t["assigned_user_id"], g.user["nome"], motivo)
        flash("Conclusão rejeitada. O chamado voltou para atendimento.", "info")
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

@bp.post("/tickets/<int:ticket_id>/steps/<int:step_id>/toggle")
@login_required
@role_required('admin', 'operador')
def toggle_step_route(ticket_id: int, step_id: int):
    done = request.form.get("done") == "on"
    try:
        toggle_step(step_id, done)
        flash("Etapa atualizada.", "success")
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

@bp.post("/tickets/<int:ticket_id>/steps/add")
@login_required
@role_required('admin', 'operador')
def add_step_route(ticket_id: int):
    text = request.form.get("text", "")
    position = request.form.get("position", "end")
    ref_step_id = request.form.get("ref_step_id")
    try:
        add_step(ticket_id, text, position=position, ref_step_id=int(ref_step_id) if ref_step_id else None)
        flash("Etapa adicionada.", "success")
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

@bp.post("/tickets/<int:ticket_id>/steps/<int:step_id>/delete")
@login_required
@role_required('admin', 'operador')
def delete_step_route(ticket_id: int, step_id: int):
    try:
        delete_step(ticket_id, step_id)
        flash("Etapa removida.", "success")
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

@bp.post("/tickets/<int:ticket_id>/steps/<int:step_id>/move")
@login_required
@role_required('admin', 'operador')
def move_step_route(ticket_id: int, step_id: int):
    direction = request.form.get("direction", "")
    try:
        move_step(ticket_id, step_id, direction)
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

@bp.post("/tickets/<int:ticket_id>/upload")
@login_required
def upload_attachment(ticket_id: int):
    t = get_ticket(ticket_id)
    if not t:
        flash("Chamado não encontrado.", "error")
        return redirect(url_for("routes.index"))

    # Solicitante só pode enviar anexos enquanto ninguém assumiu o chamado
    if not _solicitante_can_edit(t):
        flash("Este chamado já está em atendimento. Você não pode enviar anexos enquanto há um responsável atribuído.", "error")
        return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

    # Operador/admin podem sempre — mas solicitante só acessa o próprio chamado
    if g.user["role"] == "solicitante" and t["requester_user_id"] != g.user["id"]:
        flash("Você não tem permissão para enviar anexos neste chamado.", "error")
        return redirect(url_for("routes.my_tickets"))

    files = request.files.getlist("files")
    files = [f for f in files if f and (f.filename or "").strip() != ""]
    if not files:
        flash("Selecione pelo menos um arquivo.", "error")
        return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

    max_files = current_app.config["MAX_ATTACHMENTS_PER_TYPE"].get(t["tipo"], 20)
    current = count_attachments(ticket_id)
    if current + len(files) > max_files:
        flash(f"Limite de anexos excedido: {current}/{max_files} já anexados. Você tentou enviar {len(files)}.", "error")
        return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

    folder = os.path.join(current_app.config["UPLOAD_FOLDER"], str(ticket_id))
    os.makedirs(folder, exist_ok=True)

    allowed_list = ", ".join(sorted(current_app.config["ALLOWED_EXTENSIONS"]))
    saved = 0

    for file in files:
        original = secure_filename(file.filename)
        if not _allowed_file(original):
            flash(f"Arquivo bloqueado: {original}. Permitidos: {allowed_list}.", "error")
            continue

        ext = os.path.splitext(original)[1].lower()
        stored = f"{uuid.uuid4().hex}{ext}"
        save_path = os.path.join(folder, stored)
        file.save(save_path)

        add_attachment(ticket_id, stored, original, file.mimetype or "", os.path.getsize(save_path))
        saved += 1

    if saved:
        flash(f"{saved} anexo(s) enviado(s).", "success")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

@bp.post("/tickets/<int:ticket_id>/files/<int:attachment_id>/delete")
@login_required
def delete_attachment_route(ticket_id: int, attachment_id: int):
    t = get_ticket(ticket_id)
    if not t:
        flash("Chamado não encontrado.", "error")
        return redirect(url_for("routes.index"))
    # Solicitante não pode excluir anexos quando chamado está em atendimento
    if g.user["role"] == "solicitante":
        if not _solicitante_can_edit(t):
            flash("Não é possível excluir anexos enquanto o chamado está em atendimento.", "error")
            return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))
        if t["requester_user_id"] != g.user["id"]:
            flash("Você não tem permissão para excluir anexos neste chamado.", "error")
            return redirect(url_for("routes.my_tickets"))
    try:
        stored = delete_attachment(ticket_id, attachment_id)
        # remove do disco
        folder = os.path.join(current_app.config["UPLOAD_FOLDER"], str(ticket_id))
        path = os.path.join(folder, stored)
        if os.path.exists(path):
            os.remove(path)
        flash("Anexo removido.", "success")
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

@bp.get("/tickets/<int:ticket_id>/files/<int:attachment_id>")
@login_required
def download_attachment(ticket_id: int, attachment_id: int):
    row = get_db().execute(
        "SELECT stored_name, original_name FROM attachments WHERE id=? AND ticket_id=?",
        (attachment_id, ticket_id)
    ).fetchone()
    if not row:
        abort(404)
    folder = os.path.join(current_app.config["UPLOAD_FOLDER"], str(ticket_id))
    return send_from_directory(folder, row["stored_name"], as_attachment=True, download_name=row["original_name"])


@bp.get("/kanban")
@login_required
def kanban():
    # Mostra apenas não concluídos/cancelados por padrão
    db = get_db()
    stats_today = _today_ymd()

    # Defina as colunas que fazem sentido pro fluxo
    statuses = [
        ("ABERTO", "Aberto"),
        ("EM_ANDAMENTO", "Em andamento"),
        ("AGUARDANDO_FORNECEDOR", "Aguardando fornecedor"),
        ("AGUARDANDO_APROVACAO", "Aguardando aprovação"),
        ("ENVIADO", "Enviado"),
        ("CONCLUIDO", "Concluído"),
        ("CANCELADO", "Cancelado"),
    ]

    rows = db.execute("""SELECT id, tipo, titulo, solicitante, responsavel, prioridade, status, data_limite, atualizado_em
                           FROM tickets
                           ORDER BY atualizado_em DESC""").fetchall()

    by_status = {s[0]: [] for s in statuses}
    for t in rows:
        st = t["status"]
        if st not in by_status:
            by_status.setdefault(st, []).append(t)
        else:
            by_status[st].append(t)

    columns = [{"status": s, "title": title, "cards": by_status.get(s, [])} for s, title in statuses]
    return render_template("kanban.html", columns=columns, stats_today=stats_today)


@bp.get("/export.csv")
@login_required
def export_csv():
    filters = {k: request.args.get(k, "") for k in ["status","tipo","prioridade","responsavel","q","sort_by","asset_id"]}
    tickets = list_tickets(filters)
    si = StringIO()
    w = csv.writer(si)
    w.writerow(["id","tipo","titulo","solicitante","responsavel","prioridade","status","data_limite","atualizado_em"])
    for r in tickets:
        w.writerow([r["id"],r["tipo"],r["titulo"],r["solicitante"],r["responsavel"],r["prioridade"],r["status"],r["data_limite"],r["atualizado_em"]])
    return Response(si.getvalue(), mimetype="text/csv", headers={"Content-Disposition":"attachment; filename=chamados_export.csv"})

# ══════════════════════════════════════════════════════════════════════════════
# NOVAS MELHORIAS
# ══════════════════════════════════════════════════════════════════════════════

# ── Comentários ───────────────────────────────────────────────────────────────

@bp.post("/tickets/<int:ticket_id>/comentarios")
@login_required
def add_comment_route(ticket_id: int):
    t = get_ticket(ticket_id)
    if not t:
        flash("Chamado não encontrado.", "error")
        return redirect(url_for("routes.index"))
    conteudo = request.form.get("conteudo", "").strip()
    if not conteudo:
        flash("Comentário não pode estar vazio.", "error")
        return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))
    interno = request.form.get("interno") == "1" and g.user["role"] in ("admin", "operador")
    add_comment(ticket_id, g.user["id"], g.user["nome"], conteudo, interno)
    flash("Comentário adicionado.", "success")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id) + "#comentarios")

@bp.post("/tickets/<int:ticket_id>/comentarios/<int:comment_id>/excluir")
@login_required
def delete_comment_route(ticket_id: int, comment_id: int):
    delete_comment(comment_id, g.user["id"])
    flash("Comentário removido.", "success")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id) + "#comentarios")

# ── Transferência ─────────────────────────────────────────────────────────────

@bp.post("/tickets/<int:ticket_id>/transferir")
@login_required
@role_required("admin", "operador")
def transfer_ticket_route(ticket_id: int):
    t = get_ticket(ticket_id)
    if not t:
        flash("Chamado não encontrado.", "error")
        return redirect(url_for("routes.index"))
    para_user_id = request.form.get("para_user_id", "")
    motivo = request.form.get("motivo", "")
    if not para_user_id or not para_user_id.isdigit():
        flash("Selecione um operador para transferir.", "error")
        return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))
    from .services.auth_service import get_user
    dest = get_user(int(para_user_id))
    if not dest:
        flash("Operador não encontrado.", "error")
        return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))
    try:
        transfer_ticket(ticket_id, int(para_user_id), dest["nome"],
                        g.user["id"], g.user["nome"], motivo)
        on_transferencia(ticket_id, t["titulo"], int(para_user_id), g.user["nome"])
        flash(f"Chamado transferido para {dest['nome']}.", "success")
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

# ── Reabertura ────────────────────────────────────────────────────────────────

@bp.post("/tickets/<int:ticket_id>/reabrir")
@login_required
def reabrir_ticket_route(ticket_id: int):
    t = get_ticket(ticket_id)
    if not t:
        flash("Chamado não encontrado.", "error")
        return redirect(url_for("routes.index"))
    if g.user["role"] != "admin" and t["requester_user_id"] != g.user["id"]:
        flash("Apenas quem abriu o chamado ou um admin pode reabri-lo.", "error")
        return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))
    try:
        reabrir_ticket(ticket_id, g.user["nome"])
        flash("Chamado reaberto com sucesso.", "success")
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

@bp.post("/tickets/<int:ticket_id>/devolver")
@login_required
@role_required("admin", "operador")
def devolver_ticket_route(ticket_id: int):
    t = get_ticket(ticket_id)
    if not t:
        flash("Chamado não encontrado.", "error")
        return redirect(url_for("routes.index"))
    motivo = request.form.get("motivo_devolucao", "").strip()
    if not motivo:
        flash("Informe o motivo da devolução para o solicitante.", "error")
        return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))
    try:
        devolver_ao_solicitante(ticket_id, g.user["nome"], motivo)
        # Notificar solicitante
        if t["requester_user_id"]:
            from .notifications import criar_notificacao
            criar_notificacao(
                t["requester_user_id"], "AGUARDANDO_INFO",
                f"Seu chamado precisa de complemento: {t['titulo']}",
                f"{g.user['nome']} devolveu seu chamado. Motivo: {motivo[:120]}",
                ticket_id
            )
        flash("Chamado devolvido ao solicitante para complemento de informações.", "info")
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

@bp.post("/tickets/<int:ticket_id>/reenviar")
@login_required
def reenviar_ticket_route(ticket_id: int):
    t = get_ticket(ticket_id)
    if not t:
        flash("Chamado não encontrado.", "error")
        return redirect(url_for("routes.my_tickets"))
    # Só quem abriu pode reenviar
    if t["requester_user_id"] != g.user["id"] and g.user["role"] != "admin":
        flash("Apenas quem abriu o chamado pode reenviar.", "error")
        return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))
    complemento = request.form.get("complemento", "").strip()
    if not complemento:
        flash("Adicione as informações solicitadas antes de reenviar.", "error")
        return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))
    try:
        # Salvar o complemento como comentário antes de reenviar
        add_comment(ticket_id, g.user["id"], g.user["nome"],
                    f"[Complemento de informações]\n{complemento}", interno=False)
        reenviar_pelo_solicitante(ticket_id, g.user["nome"], complemento)
        # Notificar o responsável (se houver)
        if t["assigned_user_id"]:
            from .notifications import criar_notificacao
            criar_notificacao(
                t["assigned_user_id"], "CHAMADO_REENVIADO",
                f"Chamado complementado: {t['titulo']}",
                f"{g.user['nome']} adicionou as informações solicitadas e reenviou.",
                ticket_id
            )
        flash("Informações enviadas! O chamado voltou para o responsável.", "success")
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

# ── Aprovação ─────────────────────────────────────────────────────────────────

@bp.get("/aprovacoes")
@login_required
@role_required("admin")
def aprovacoes():
    db = get_db()
    pendentes = db.execute(
        """SELECT t.*, c.nome as categoria_nome, c.cor as categoria_cor
           FROM tickets t LEFT JOIN categories c ON c.id=t.categoria_id
           WHERE t.status='AGUARDANDO_APROVACAO'
           ORDER BY t.criado_em ASC"""
    ).fetchall()
    return render_template("aprovacoes.html", tickets=pendentes)

@bp.post("/tickets/<int:ticket_id>/aprovar")
@login_required
@role_required("admin")
def aprovar_ticket_route(ticket_id: int):
    t = get_ticket(ticket_id)
    if not t:
        flash("Chamado não encontrado.", "error")
        return redirect(url_for("routes.aprovacoes"))
    try:
        aprovar_ticket(ticket_id, g.user["id"], g.user["nome"])
        if t["requester_user_id"]:
            on_chamado_aprovado(ticket_id, t["titulo"], t["requester_user_id"], g.user["nome"])
        flash("Chamado aprovado e enviado para a fila.", "success")
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.aprovacoes"))

@bp.post("/tickets/<int:ticket_id>/reprovar")
@login_required
@role_required("admin")
def reprovar_ticket_route(ticket_id: int):
    t = get_ticket(ticket_id)
    if not t:
        flash("Chamado não encontrado.", "error")
        return redirect(url_for("routes.aprovacoes"))
    motivo = request.form.get("motivo", "")
    try:
        reprovar_ticket(ticket_id, g.user["nome"], motivo)
        if t["requester_user_id"]:
            on_chamado_reprovado(ticket_id, t["titulo"], t["requester_user_id"], g.user["nome"])
        flash("Chamado reprovado.", "info")
    except Exception as e:
        flash(str(e), "error")
    return redirect(url_for("routes.aprovacoes"))

# ── TMA ───────────────────────────────────────────────────────────────────────

@bp.get("/api/tma")
@login_required
def api_tma():
    uid = request.args.get("user_id")
    cid = request.args.get("categoria_id")
    stats = tma_stats(
        user_id=int(uid) if uid and uid.isdigit() else None,
        categoria_id=int(cid) if cid and cid.isdigit() else None,
    )
    return jsonify(stats)

# ── Busca avançada ────────────────────────────────────────────────────────────

@bp.get("/busca-avancada")
@login_required
@role_required("admin", "operador")
def advanced_search():
    f = {k: request.args.get(k, "") for k in
         ["q","status","categoria_id","classificacao","responsavel","data_inicio","data_fim"]}
    results = []
    if any(f.values()):
        results = search_tickets_advanced(
            **f,
            user_id=g.user["id"],
            user_role=g.user["role"],
        )
    categories = list_categories(only_active=True)
    return render_template("advanced_search.html",
                           filters=f, results=results,
                           STATUSES=STATUSES, CLASSIFICATIONS=CLASSIFICATIONS,
                           categories=categories)

# ── Logs / Auditoria ──────────────────────────────────────────────────────────

@bp.get("/logs")
@login_required
@role_required("admin")
def logs_audit():
    db = get_db()
    filtros, where_sql, params = _build_logs_filters_and_query(request.args)
    logs = _fetch_logs_for_audit(db, where_sql, params, limit=500)

    eventos = db.execute("SELECT DISTINCT evento FROM ticket_log ORDER BY evento ASC").fetchall()

    total = db.execute(f"""
        SELECT COUNT(*) AS total
          FROM ticket_log tl
          LEFT JOIN tickets t ON t.id = tl.ticket_id
          {where_sql}
    """, params).fetchone()["total"]

    return render_template("logs.html", logs=logs, filtros=filtros, eventos=eventos, total=total)


@bp.get("/logs/exportar/excel")
@login_required
@role_required("admin")
def export_logs_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    db = get_db()
    filtros, where_sql, params = _build_logs_filters_and_query(request.args)
    logs = _fetch_logs_for_audit(db, where_sql, params, limit=5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Logs"
    ws.freeze_panes = "A5"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    sub_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="D9D9D9")

    ws["A1"] = "Relatório de Auditoria de Logs"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A3"] = (
        f"Filtros | Busca: {filtros['q'] or '-'} | Evento: {filtros['evento'] or '-'} | "
        f"Chamado: {filtros['ticket_id'] or '-'} | Início: {filtros['data_inicio'] or '-'} | Fim: {filtros['data_fim'] or '-'}"
    )

    headers = ["Data/Hora", "ID Chamado", "Número do Chamado", "Título", "Evento", "Detalhe", "Status Atual"]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, log in enumerate(logs, start=5):
        values = [
            log["criado_em"],
            log["ticket_id"],
            log["numero_chamado"],
            log["titulo"],
            log["evento"],
            log["detalhe"],
            log["status"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value or "-")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin_gray)
        if row_idx % 2 == 0:
            for col_idx in range(1, 8):
                ws.cell(row=row_idx, column=col_idx).fill = sub_fill

    widths = {"A": 20, "B": 12, "C": 18, "D": 34, "E": 22, "F": 60, "G": 16}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[4].height = 24

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=_logs_export_filename("auditoria_logs", "xlsx"),
    )


@bp.get("/logs/exportar/pdf")
@login_required
@role_required("admin")
def export_logs_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    db = get_db()
    filtros, where_sql, params = _build_logs_filters_and_query(request.args)
    logs = _fetch_logs_for_audit(db, where_sql, params, limit=1500)

    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontName = "Helvetica-Bold"
    title_style.fontSize = 16
    title_style.textColor = colors.HexColor("#1F4E78")

    meta_style = ParagraphStyle(
        "AuditMeta",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#374151"),
    )
    cell_style = ParagraphStyle(
        "AuditCell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9,
    )
    cell_style_center = ParagraphStyle(
        "AuditCellCenter",
        parent=cell_style,
        alignment=1,
    )

    story = [
        Paragraph("Relatório de Auditoria de Logs", title_style),
        Spacer(1, 4),
        Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", meta_style),
        Paragraph(
            f"Filtros aplicados: busca={filtros['q'] or '-'} | evento={filtros['evento'] or '-'} | chamado={filtros['ticket_id'] or '-'} | período={filtros['data_inicio'] or '-'} até {filtros['data_fim'] or '-'}",
            meta_style,
        ),
        Paragraph(
            "Formato de apoio à auditoria. A exportação em PDF leva até 1.500 registros filtrados para preservar legibilidade e desempenho.",
            meta_style,
        ),
        Spacer(1, 8),
    ]

    table_data = [[
        Paragraph("Data/Hora", cell_style_center),
        Paragraph("Chamado", cell_style_center),
        Paragraph("Evento", cell_style_center),
        Paragraph("Detalhe", cell_style_center),
        Paragraph("Status", cell_style_center),
    ]]

    for log in logs:
        chamado = log["numero_chamado"] or (f"#{log['ticket_id']}" if log["ticket_id"] else "-")
        if log["titulo"]:
            chamado = f"{chamado}<br/>{log['titulo']}"
        table_data.append([
            Paragraph(log["criado_em"] or "-", cell_style),
            Paragraph(chamado, cell_style),
            Paragraph(log["evento"] or "-", cell_style),
            Paragraph((log["detalhe"] or "-").replace("\n", "<br/>"), cell_style),
            Paragraph(log["status"] or "-", cell_style),
        ])

    table = Table(table_data, colWidths=[34*mm, 42*mm, 34*mm, 118*mm, 28*mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#EEF5FB")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    story.append(table)

    def _add_page_number(canvas, _doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#4B5563"))
        canvas.drawRightString(285 * mm, 8 * mm, f"Página {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_add_page_number, onLaterPages=_add_page_number)
    bio.seek(0)

    return send_file(
        bio,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=_logs_export_filename("auditoria_logs", "pdf"),
    )

# ── Grupos de Operadores ──────────────────────────────────────────────────────

@bp.get("/grupos")
@login_required
@role_required("admin")
def groups_list():
    from .services.user_service import list_users

    groups = list_groups()
    operators = [u for u in list_users() if u["role"] in ("admin", "operador")]
    categories = list_categories()
    group_members_map = {g_["id"]: get_group_members(g_["id"]) for g_ in groups}
    group_categories_map = {g_["id"]: get_group_categories(g_["id"]) for g_ in groups}
    return render_template(
        "groups.html",
        groups=groups,
        operators=operators,
        categories=categories,
        group_members_map=group_members_map,
        group_categories_map=group_categories_map,
    )

@bp.post("/grupos/novo")
@login_required
@role_required("admin")
def create_group_route():
    nome = request.form.get("nome","").strip()
    if not nome:
        flash("Nome é obrigatório.", "error")
        return redirect(url_for("routes.groups_list"))
    gid = create_group(nome, request.form.get("descricao",""), request.form.get("cor","#6366f1"))
    members = request.form.getlist("user_ids")
    cats = request.form.getlist("category_ids")
    if members:
        set_group_members(gid, [int(x) for x in members if x.isdigit()])
    if cats:
        set_group_categories(gid, [int(x) for x in cats if x.isdigit()])
    flash(f"Grupo '{nome}' criado.", "success")
    return redirect(url_for("routes.groups_list"))

@bp.post("/grupos/<int:group_id>/editar")
@login_required
@role_required("admin")
def edit_group_route(group_id: int):
    update_group(group_id, request.form.get("nome",""), request.form.get("descricao",""),
                 request.form.get("cor","#6366f1"), request.form.get("ativo")=="1")
    members = request.form.getlist("user_ids")
    cats = request.form.getlist("category_ids")
    set_group_members(group_id, [int(x) for x in members if x.isdigit()])
    set_group_categories(group_id, [int(x) for x in cats if x.isdigit()])
    flash("Grupo atualizado.", "success")
    return redirect(url_for("routes.groups_list"))

@bp.post("/tickets/<int:ticket_id>/atribuir-grupo")
@login_required
@role_required("admin", "operador")
def assign_group_route(ticket_id: int):
    group_id = request.form.get("group_id","")
    if not group_id or not group_id.isdigit():
        flash("Selecione um grupo.", "error")
        return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))
    result = assign_ticket_to_group(ticket_id, int(group_id))
    if result:
        flash(f"Chamado atribuído a {result['nome']} (menor carga do grupo).", "success")
    else:
        flash("Grupo sem membros ativos.", "error")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

# ── Webhooks ──────────────────────────────────────────────────────────────────

@bp.get("/admin/webhooks")
@login_required
@role_required("admin")
def webhooks_list():
    hooks = list_webhooks()
    WEBHOOK_EVENTS = [
        "ticket.criado", "ticket.assumido", "ticket.status_alterado",
        "ticket.finalizado", "ticket.concluido", "ticket.transferido",
        "ticket.aprovado", "ticket.reprovado",
    ]
    return render_template("webhooks.html", hooks=hooks, WEBHOOK_EVENTS=WEBHOOK_EVENTS)

@bp.post("/admin/webhooks/novo")
@login_required
@role_required("admin")
def create_webhook_route():
    nome = request.form.get("nome","").strip()
    url_hook = request.form.get("url","").strip()
    eventos = request.form.getlist("eventos")
    secret = request.form.get("secret","").strip()
    if not nome or not url_hook:
        flash("Nome e URL são obrigatórios.", "error")
        return redirect(url_for("routes.webhooks_list"))
    create_webhook(nome, url_hook, eventos, secret)
    flash(f"Webhook '{nome}' criado.", "success")
    return redirect(url_for("routes.webhooks_list"))

@bp.post("/admin/webhooks/<int:hook_id>/editar")
@login_required
@role_required("admin")
def edit_webhook_route(hook_id: int):
    update_webhook(hook_id, request.form.get("nome",""), request.form.get("url",""),
                   request.form.getlist("eventos"), request.form.get("ativo")=="1",
                   request.form.get("secret",""))
    flash("Webhook atualizado.", "success")
    return redirect(url_for("routes.webhooks_list"))

@bp.post("/admin/webhooks/<int:hook_id>/excluir")
@login_required
@role_required("admin")
def delete_webhook_route(hook_id: int):
    delete_webhook(hook_id)
    flash("Webhook removido.", "success")
    return redirect(url_for("routes.webhooks_list"))

# ══════════════════════════════════════════════════════════════════════════════
# CHAMADOS RECORRENTES
# ══════════════════════════════════════════════════════════════════════════════

@bp.get("/recorrentes")
@login_required
@role_required("admin")
def recurring_list():
    db = get_db()
    schedules = db.execute("SELECT * FROM recurring_tickets ORDER BY titulo").fetchall()
    categories = list_categories(only_active=True)
    return render_template("recurring.html", schedules=schedules,
                            categories=categories, PRIORITIES=PRIORITIES)

@bp.post("/recorrentes/novo")
@login_required
@role_required("admin")
def recurring_create():
    import json as _json
    titulo = request.form.get("titulo","").strip()
    if not titulo:
        flash("Título é obrigatório.", "error")
        return redirect(url_for("routes.recurring_list"))
    ticket_data = {
        "descricao": request.form.get("descricao",""),
        "prioridade": request.form.get("prioridade","MEDIA"),
        "categoria_id": request.form.get("categoria_id",""),
        "classificacao": request.form.get("classificacao","REQUISICAO"),
    }
    db = get_db()
    db.execute(
        """INSERT INTO recurring_tickets (titulo, frequencia, dia_execucao, hora_execucao,
           ticket_data, ativo, criado_em) VALUES (?,?,?,?,?,1,?)""",
        (titulo, request.form.get("frequencia","mensal"),
         request.form.get("dia_execucao") or None,
         request.form.get("hora_execucao","08:00"),
         _json.dumps(ticket_data), datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    )
    db.commit()
    flash(f"Chamado recorrente '{titulo}' criado.", "success")
    return redirect(url_for("routes.recurring_list"))

@bp.post("/recorrentes/<int:rid>/toggle")
@login_required
@role_required("admin")
def recurring_toggle(rid: int):
    db = get_db()
    row = db.execute("SELECT ativo FROM recurring_tickets WHERE id=?", (rid,)).fetchone()
    if row:
        db.execute("UPDATE recurring_tickets SET ativo=? WHERE id=?", (0 if row["ativo"] else 1, rid))
        db.commit()
    return redirect(url_for("routes.recurring_list"))

@bp.post("/recorrentes/<int:rid>/excluir")
@login_required
@role_required("admin")
def recurring_delete(rid: int):
    get_db().execute("DELETE FROM recurring_tickets WHERE id=?", (rid,))
    get_db().commit()
    flash("Agendamento removido.", "success")
    return redirect(url_for("routes.recurring_list"))

# ══════════════════════════════════════════════════════════════════════════════
# PORTAL EXTERNO — geração de token
# ══════════════════════════════════════════════════════════════════════════════

@bp.post("/tickets/<int:ticket_id>/gerar-token")
@login_required
@role_required("admin", "operador")
def gerar_portal_token(ticket_id: int):
    t = get_ticket(ticket_id)
    if not t:
        flash("Chamado não encontrado.", "error")
        return redirect(url_for("routes.index"))
    from .portal import create_portal_token
    email = request.form.get("email","").strip()
    token = create_portal_token(ticket_id, email=email, expira_horas=720)
    app_url = request.host_url.rstrip("/")
    link = f"{app_url}/portal/{token}"
    flash(f"Link gerado: {link}", "success")
    # Envia por e-mail se fornecido
    if email:
        from .notify import notify_async
        from flask import current_app
        body = (
            f"Olá,\n\n"
            f"Você pode acompanhar seu chamado pelo link abaixo:\n\n"
            f"  {link}\n\n"
            f"Chamado: {t['titulo']}\n"
            f"Número : {t['numero_chamado'] or '#'+str(ticket_id)}\n\n"
            f"O link é válido por 30 dias."
        )
        notify_async(dict(current_app.config), [email],
                     f"[CCTI] Acompanhe seu chamado {t['numero_chamado'] or '#'+str(ticket_id)}", body)
        flash(f"Link enviado para {email}.", "info")
    return redirect(url_for("routes.ticket_detail", ticket_id=ticket_id))

# ══════════════════════════════════════════════════════════════════════════════
# BUSCA GLOBAL ⌘K
# ══════════════════════════════════════════════════════════════════════════════

@bp.get("/api/search")
@login_required
def global_search_api():
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify(results=[])

    db = get_db()
    like = f"%{q}%"
    results = []

    # Chamados
    uid  = g.user["id"]
    role = g.user["role"]
    scope = ""
    scope_params = []
    if role == "operador":
        allowed = db.execute("SELECT category_id FROM user_categories WHERE user_id=?", (uid,)).fetchall()
        ids = [r["category_id"] for r in allowed]
        if ids:
            scope = f" AND t.categoria_id IN ({','.join('?'*len(ids))})"
            scope_params = ids

    ticket_rows = db.execute(
        f"""SELECT t.id, t.numero_chamado, t.titulo, t.status, t.prioridade
            FROM tickets t WHERE (t.titulo LIKE ? OR t.numero_chamado LIKE ? OR t.solicitante LIKE ?){scope}
            ORDER BY t.atualizado_em DESC LIMIT 8""",
        [like, like, like] + scope_params
    ).fetchall()
    for r in ticket_rows:
        results.append({
            "type": "ticket",
            "id": r["id"],
            "label": f"{r['numero_chamado'] or '#'+str(r['id'])} — {r['titulo']}",
            "sub": f"{r['status'].replace('_',' ')} · {r['prioridade']}",
            "url": url_for("routes.ticket_detail", ticket_id=r["id"]),
        })

    # KB — artigos
    kb_rows = db.execute(
        "SELECT id, titulo FROM kb_articles WHERE (titulo LIKE ? OR tags LIKE ? OR conteudo LIKE ?) AND publico=1 LIMIT 4",
        [like, like, like]
    ).fetchall()
    for r in kb_rows:
        results.append({
            "type": "kb",
            "id": r["id"],
            "label": r["titulo"],
            "sub": "Base de conhecimento",
            "url": url_for("kb.kb_article", article_id=r["id"]),
        })

    # Usuários (só admin)
    if role == "admin":
        user_rows = db.execute(
            "SELECT id, nome, email, role FROM users WHERE nome LIKE ? OR email LIKE ? LIMIT 4",
            [like, like]
        ).fetchall()
        for r in user_rows:
            results.append({
                "type": "user",
                "id": r["id"],
                "label": r["nome"],
                "sub": f"{r['email']} · {r['role']}",
                "url": url_for("admin.users"),
            })

    return jsonify(results=results)
