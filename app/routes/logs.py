"""Rotas de Auditoria de Logs."""

from io import BytesIO
from datetime import datetime
from flask import render_template, request, g, send_file

from app.routes import bp
from app.auth.decorators import login_required, role_required
from app.db import get_db


def _build_logs_filters_and_query(args):
    filtros = {
        "q": (args.get("q") or "").strip(),
        "evento": (args.get("evento") or "").strip(),
        "ticket_id": (args.get("ticket_id") or "").strip(),
        "data_inicio": (args.get("data_inicio") or "").strip(),
        "data_fim": (args.get("data_fim") or "").strip(),
    }
    where = []
    params = []
    if filtros["ticket_id"].isdigit():
        where.append("tl.ticket_id = ?"); params.append(int(filtros["ticket_id"]))
    if filtros["evento"]:
        where.append("tl.evento = ?"); params.append(filtros["evento"].upper())
    if filtros["q"]:
        like = f"%{filtros['q']}%"
        where.append("(tl.detalhe LIKE ? OR tl.evento LIKE ? OR t.titulo LIKE ? OR t.numero_chamado LIKE ?)")
        params.extend([like, like, like, like])
    if filtros["data_inicio"]:
        where.append("date(tl.criado_em) >= date(?)"); params.append(filtros["data_inicio"])
    if filtros["data_fim"]:
        where.append("date(tl.criado_em) <= date(?)"); params.append(filtros["data_fim"])
    where_sql = f"WHERE {' AND '.join(where)}" if where else ""
    return filtros, where_sql, params


def _fetch_logs_for_audit(db, where_sql, params, limit=500):
    sql = f"""SELECT tl.id, tl.ticket_id, tl.evento, tl.detalhe, tl.criado_em,
               t.numero_chamado, t.titulo, t.status
          FROM ticket_log tl LEFT JOIN tickets t ON t.id = tl.ticket_id {where_sql}
         ORDER BY tl.criado_em DESC, tl.id DESC"""
    if limit:
        sql += f" LIMIT {int(limit)}"
    return db.execute(sql, params).fetchall()


def _logs_export_filename(prefix, ext):
    return f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"


@bp.get("/logs")
@login_required
@role_required("admin")
def logs_audit():
    db = get_db()
    filtros, where_sql, params = _build_logs_filters_and_query(request.args)
    logs = _fetch_logs_for_audit(db, where_sql, params, limit=500)
    eventos = db.execute("SELECT DISTINCT evento FROM ticket_log ORDER BY evento ASC").fetchall()
    total = db.execute(f"SELECT COUNT(*) AS total FROM ticket_log tl LEFT JOIN tickets t ON t.id = tl.ticket_id {where_sql}", params).fetchone()["total"]
    return render_template("logs.html", logs=logs, filtros=filtros, eventos=eventos, total=total)


@bp.get("/logs/exportar/excel")
@login_required
@role_required("admin")
def export_logs_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    db = get_db()
    filtros, where_sql, params = _build_logs_filters_and_query(request.args)
    logs = _fetch_logs_for_audit(db, where_sql, params, limit=5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Logs"
    ws.freeze_panes = "A5"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    sub_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="D9D9D9")

    ws["A1"] = "Relatório de Auditoria de Logs"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A3"] = (f"Filtros | Busca: {filtros['q'] or '-'} | Evento: {filtros['evento'] or '-'} | "
                f"Chamado: {filtros['ticket_id'] or '-'} | Início: {filtros['data_inicio'] or '-'} | Fim: {filtros['data_fim'] or '-'}")

    headers = ["Data/Hora", "ID Chamado", "Número do Chamado", "Título", "Evento", "Detalhe", "Status Atual"]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, log in enumerate(logs, start=5):
        values = [log["criado_em"], log["ticket_id"], log["numero_chamado"], log["titulo"],
                  log["evento"], log["detalhe"], log["status"]]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value or "-")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin_gray)
        if row_idx % 2 == 0:
            for col_idx in range(1, 8):
                ws.cell(row=row_idx, column=col_idx).fill = sub_fill

    for col_letter, width in {"A": 20, "B": 12, "C": 18, "D": 34, "E": 22, "F": 60, "G": 16}.items():
        ws.column_dimensions[col_letter].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=_logs_export_filename("auditoria_logs", "xlsx"))


@bp.get("/logs/exportar/pdf")
@login_required
@role_required("admin")
def export_logs_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    db = get_db()
    filtros, where_sql, params = _build_logs_filters_and_query(request.args)
    logs = _fetch_logs_for_audit(db, where_sql, params, limit=1500)

    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(A4), leftMargin=12*mm, rightMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontName = "Helvetica-Bold"; title_style.fontSize = 16
    title_style.textColor = colors.HexColor("#1F4E78")
    meta_style = ParagraphStyle("AuditMeta", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.5, leading=11, textColor=colors.HexColor("#374151"))
    cell_style = ParagraphStyle("AuditCell", parent=styles["BodyText"], fontName="Helvetica", fontSize=7.5, leading=9)
    cell_style_center = ParagraphStyle("AuditCellCenter", parent=cell_style, alignment=1)

    story = [
        Paragraph("Relatório de Auditoria de Logs", title_style), Spacer(1, 4),
        Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", meta_style),
        Paragraph(f"Filtros: busca={filtros['q'] or '-'} | evento={filtros['evento'] or '-'} | chamado={filtros['ticket_id'] or '-'} | período={filtros['data_inicio'] or '-'} até {filtros['data_fim'] or '-'}", meta_style),
        Spacer(1, 8),
    ]
    table_data = [[Paragraph(h, cell_style_center) for h in ["Data/Hora", "Chamado", "Evento", "Detalhe", "Status"]]]
    for log in logs:
        chamado = log["numero_chamado"] or (f"#{log['ticket_id']}" if log["ticket_id"] else "-")
        if log["titulo"]: chamado = f"{chamado}<br/>{log['titulo']}"
        table_data.append([
            Paragraph(log["criado_em"] or "-", cell_style),
            Paragraph(chamado, cell_style),
            Paragraph(log["evento"] or "-", cell_style),
            Paragraph((log["detalhe"] or "-").replace("\n", "<br/>"), cell_style),
            Paragraph(log["status"] or "-", cell_style),
        ])
    table = Table(table_data, colWidths=[34*mm, 42*mm, 34*mm, 118*mm, 28*mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#EEF5FB")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)

    def _page_num(canvas, _doc):
        canvas.saveState(); canvas.setFont("Helvetica", 8); canvas.setFillColor(colors.HexColor("#4B5563"))
        canvas.drawRightString(285*mm, 8*mm, f"Página {canvas.getPageNumber()}"); canvas.restoreState()
    doc.build(story, onFirstPage=_page_num, onLaterPages=_page_num)
    bio.seek(0)
    return send_file(bio, mimetype="application/pdf", as_attachment=True, download_name=_logs_export_filename("auditoria_logs", "pdf"))
