from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from app.models import ticket_report_metrics
from app.services.asset_service import list_assets, asset_dashboard

def _header_style(ws, row=1):
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[row]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

def build_tickets_xlsx(start_date=None, end_date=None, sla_days=3) -> bytes:
    data = ticket_report_metrics(start_date, end_date, sla_days)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumo"
    ws1.append(["Métrica", "Valor"])
    rows = [
        ["Período inicial", start_date or "-"],
        ["Período final", end_date or "-"],
        ["SLA (dias)", sla_days],
        ["Total de chamados", data["total"]],
        ["Chamados resolvidos", data["resolved_total"]],
        ["Tempo médio de resolução (dias)", data["avg_resolution_days"]],
        ["Dentro do SLA", data["within_sla_count"]],
        ["Fora do SLA", data["outside_sla_count"]],
        ["% dentro do SLA", data["within_sla_pct"]],
    ]
    for r in rows: ws1.append(r)
    _header_style(ws1)
    ws2 = wb.create_sheet("Chamados")
    ws2.append(["ID","Tipo","Título","Solicitante","Responsável","Prioridade","Status","Data limite","Criado em","Encerrado em","Tempo resolução (dias)","Dentro SLA"])
    for item in data["items"]:
        ws2.append([item["id"], item["tipo"], item["titulo"], item["solicitante"], item["responsavel"], item["prioridade"], item["status"], item["data_limite"], item["criado_em"], item.get("closed_em"), item["resolution_days"], "SIM" if item["within_sla"] is True else ("NÃO" if item["within_sla"] is False else "-")])
    _header_style(ws2)
    for ws in wb.worksheets:
        for col in ws.columns:
            letter = col[0].column_letter
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[letter].width = min(max_len + 2, 35)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def build_assets_xlsx() -> bytes:
    stats = asset_dashboard()
    items = [dict(r) for r in list_assets({})]
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumo"
    ws1.append(["Métrica","Valor"])
    ws1.append(["Total de ativos", stats["total"]])
    for k,v in stats["by_status"].items(): ws1.append([f"Status: {k}", v])
    for k,v in stats["by_tipo"].items(): ws1.append([f"Tipo: {k}", v])
    _header_style(ws1)
    ws2 = wb.create_sheet("Ativos")
    ws2.append(["ID","Tag","Tipo","Modelo","Serial","Base/Local","Responsável","Status","Atualizado em"])
    for a in items:
        ws2.append([a["id"], a["tag"], a["tipo"], a["modelo"], a["serial_number"], a["local_base"], a["responsavel"], a["status"], a["atualizado_em"]])
    _header_style(ws2)
    for ws in wb.worksheets:
        for col in ws.columns:
            letter = col[0].column_letter
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[letter].width = min(max_len + 2, 35)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def build_tickets_pdf(start_date=None, end_date=None, sla_days=3) -> bytes:
    data = ticket_report_metrics(start_date, end_date, sla_days)
    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [Paragraph("Relatório de Chamados e SLA", styles["Title"]),
             Paragraph(f"Período: {start_date or '-'} até {end_date or '-'} | SLA: {sla_days} dia(s)", styles["Normal"]),
             Spacer(1,12)]
    summary = [["Métrica","Valor"],["Total de chamados",str(data["total"])],["Resolvidos",str(data["resolved_total"])],["Tempo médio (dias)",str(data["avg_resolution_days"])],["Dentro SLA",str(data["within_sla_count"])],["Fora SLA",str(data["outside_sla_count"])],["% dentro SLA",f"{data['within_sla_pct']}%"]]
    t = Table(summary, hAlign="LEFT")
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E78")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),0.5,colors.grey)]))
    story.extend([t, Spacer(1,12)])
    rows = [["ID","Tipo","Título","Responsável","Status","Resolução(dias)","SLA"]]
    for item in data["items"][:20]:
        rows.append([str(item["id"]), item["tipo"], item["titulo"][:38], item["responsavel"] or "-", item["status"], "-" if item["resolution_days"] is None else str(item["resolution_days"]), "-" if item["within_sla"] is None else ("SIM" if item["within_sla"] else "NÃO")])
    t2 = Table(rows, repeatRows=1)
    t2.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E78")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),0.4,colors.grey)]))
    story.append(t2)
    doc.build(story)
    bio.seek(0)
    return bio.getvalue()

def build_assets_pdf() -> bytes:
    stats = asset_dashboard()
    items = [dict(r) for r in list_assets({})]
    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [Paragraph("Relatório de Ativos", styles["Title"]), Spacer(1,12)]
    summary = [["Métrica","Valor"],["Total de ativos",str(stats["total"])]]
    for k,v in stats["by_status"].items(): summary.append([f"Status: {k}", str(v)])
    t = Table(summary, hAlign="LEFT")
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E78")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),0.5,colors.grey)]))
    story.extend([t, Spacer(1,12)])
    rows = [["ID","Tag","Tipo","Modelo","Serial","Local","Responsável","Status"]]
    for a in items[:25]:
        rows.append([str(a["id"]), a["tag"], a["tipo"], a["modelo"][:28], (a["serial_number"] or "-")[:18], (a["local_base"] or "-")[:18], (a["responsavel"] or "-")[:18], a["status"]])
    t2 = Table(rows, repeatRows=1)
    t2.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E78")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),0.4,colors.grey)]))
    story.append(t2)
    doc.build(story)
    bio.seek(0)
    return bio.getvalue()
